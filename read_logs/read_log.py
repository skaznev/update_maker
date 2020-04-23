
#---------- ПОДКЛЮЧАЕМ БИБЛИОТЕКИ ----------

import os
import shutil
import re
import io
import openpyxl

#---------- КОНЕЦ ПОДКЛЮЧЕНИЯ БИБЛИОТЕК ----------


#---------- ОБЪЯВЛЯЕМ ПЕРЕМЕННЫЕ ----------

max_min     = 0                                    # Максимально допустимое время скрипта (минуты)
path_app    = os.getcwd()                           # Текущая директория
i           = 2                                     # Номер строки в экселе в которую пишем

#---------- КОНЕЦ ОБЪЯВЛЕНИЯ ПЕРЕМЕННЫХ ----------


#---------- РАБОТАЕМ С КАТАЛОГАМИ (СОЗДАЕМ ДИРЕКТОРИЮ В КОТОРУЮ НАЧИТЫВАЕМ ОБЪЕКТЫ С БАЗЫ) ----------

wb = openpyxl.load_workbook(filename = 'C:/projects/update_maker/logs.xlsx')
sheet = wb['test']
#sheet['A1'] = 'hui'
#wb.save('logs.xlsx')

#---------- КОНЕЦ РАБОТЫ С ДИРЕКТОРИЯМИ ----------


#---------- ОБЪЯВЛЕНИЕ ФУНКЦИИ "РАЗБОР ЛОГА" ----------

def push (object = ''):
    global i 
    text        = ''
    text_real   = ''
    real        = ''
    real_num    = 0
    print('обработка файла ' + object)
    with io.open(object, 'r', encoding = "windows-1251",errors='ignore') as log:
        for line in log:
            if r'real:' in line:
                real = re.split(r' real: ', line)
                real_num = int(real[1])
                if 1==1: #real_num/60000 > max_min:
                    text += line
                    text_real += text + '\n СКРИПТ ПРОГОНЯЛСЯ: ' + str(int(real_num/60000)) + ' МИН.\n'
                    sheet.cell(row = i, column = 2).value = real_num/60000
                    try:
                        sheet.cell(row = i, column = 1).value = text
                    except:
                        print('ошибка real: ' + str(real_num/60000))
                        i +=1
                        text = ''
                        continue
                    i +=1
                text = ''
            else:
                text += line
        # real_log = open(os.path.join(path_app , 'Dismantled_' + object) ,  "w", encoding = "windows-1251")
        # real_log.write(text_real)
        # real_log.close

#---------- КОНЕЦ ФУНКЦИИ "МЕТНИСЬ С ОБЪЕКТОМ" ----------


#---------- ТЕЛО ПРОГРАММЫ - ЧИТАЕМ КАТАЛОГ И ДЕРГАЕМСЯ ПО ОБЪЕКТАМ ----------


files = os.listdir(path_app)
all_files = [f for f in files if ( f.upper()[-4:] in [".LOG"])]
for file in all_files:
    push(file)
wb.save('logs.xlsx')

#end = input('Обработка завершена')

#---------- КОНЕЦ ТЕЛА ПРОГРАММЫ ----------
