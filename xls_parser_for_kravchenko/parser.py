
#---------- ПОДКЛЮЧАЕМ БИБЛИОТЕКИ ----------

import os
import openpyxl
import datetime

#---------- КОНЕЦ ПОДКЛЮЧЕНИЯ БИБЛИОТЕК ----------


#---------- ОБЪЯВЛЯЕМ ПЕРЕМЕННЫЕ ----------

i_row       = 1                 # Номер строки в экселе c которой начинаем читать
i_column    = 1                 # Номер колонки в экселе с которой начинаем читать
start_str   = '('               # Разделитель обозначающий начало строки
end_str     = ')'               # Разделитель обозначающий конец строки
com         = ','               # Разделитель значений в строке
string      = ''                # Итоговая строка которую напишем в файл
filename    = 'Result.txt'      # Название выходного файла

#---------- КОНЕЦ ОБЪЯВЛЕНИЯ ПЕРЕМЕННЫХ ----------

#---------- ОБЪЯВЛЕНИЕ ВСПОМОГАТЕЛЬНОЙ ФУНКЦИИ ДЛЯ ЗНАЧЕНИЙ ----------

def nvl(a):
    if a is None:
        return ' '
    elif isinstance(a,datetime.datetime):
        a = a.strftime('%d.%m.%Y')
    return str(a)

#---------- КОНЕЦ ВСПОМОГАТЕЛЬНОЙ ФУНКЦИИ ----------

#---------- ЧИТАЕМ ДИРЕКТОРИЮ И РАБОТАЕМ С ФАЙЛОМ ----------

path_book   = input('Введите полный адрес до файла Excel')
excel_book  = openpyxl.load_workbook(filename = path_book)
sheet       = excel_book.active
max_row     = sheet.max_row
max_column  = sheet.max_column

for row in range(i_row, max_row + 1):
    string += start_str
    for column in range(i_column,max_column + 1):
        string += nvl(sheet.cell(row=row,column=column).value)
        if not column == max_column:
            string += com
    string += end_str

with open(os.path.join(os.path.dirname(path_book), filename), 'w') as file:
    file.write(string)

#---------- КОНЕЦ РАБОТЫ С ФАЙЛОМ ----------