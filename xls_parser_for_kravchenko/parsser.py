
# ---------- ПОДКЛЮЧАЕМ БИБЛИОТЕКИ ----------

import os
import openpyxl
import datetime

# ---------- КОНЕЦ ПОДКЛЮЧЕНИЯ БИБЛИОТЕК ----------


# ---------- ОБЪЯВЛЯЕМ ПЕРЕМЕННЫЕ ----------

i_row = 1       # Номер строки в экселе c которой начинаем читать
i_column = 1    # Номер колонки в экселе с которой начинаем читать
start_str = '(' # Разделитель обозначающий начало строки
end_str = ')'   # Разделитель обозначающий конец строки
com = ','       # Разделитель значений в строке
string = ''     # Итоговая строка которую напишем в файл

# ---------- КОНЕЦ ОБЪЯВЛЕНИЯ ПЕРЕМЕННЫХ ----------

# ---------- ОБЪЯВЛЕНИЕ ВСПОМОГАТЕЛЬНОЙ ФУНКЦИИ ДЛЯ ЗНАЧЕНИЙ ----------

def nvl(a):
    if a is None:
        return ' '
    elif isinstance(a, datetime.datetime):
        a = a.strftime('%d.%m.%Y')
    return str(a)

# ---------- КОНЕЦ ВСПОМОГАТЕЛЬНОЙ ФУНКЦИИ ----------

# ---------- ЧИТАЕМ ДИРЕКТОРИЮ И РАБОТАЕМ С ФАЙЛОМ ----------


def execute(PATH_BOOK, SQL_TEXT):
    global i_row
    global i_column
    global start_str 
    global end_str
    global com
    global string
    global filename
    path_book   = PATH_BOOK
    sql_text    = SQL_TEXT

    excel_book  = openpyxl.load_workbook(filename=path_book)
    sheet       = excel_book.active
    max_row     = sheet.max_row

    for row in range(i_row, max_row + 1):
        if sheet.cell(row=row, column=1).value == 'Наименование':
            continue
        string += start_str
        for column in (3, 6, 10, 11, 7, 9, 2):  # (3 'C',6 'F',10 'J',11 'K',7 'G',9 'J',2 'B')
            string += nvl(sheet.cell(row=row, column=column).value)
            string += com
        string += end_str
    with open(os.path.join(os.path.dirname(path_book), os.path.split(path_book)[1][:-5] + '_SCRIPT.sql'), 'w', encoding='windows-1251') as file:
        file.write(sql_text.replace('$CLOB$', string))
    string = ''

# ---------- КОНЕЦ РАБОТЫ С ФАЙЛОМ ----------
